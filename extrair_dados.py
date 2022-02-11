from email import message
from posixpath import dirname
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
import pyautogui as py
import time
from openpyxl import Workbook, load_workbook
import os, sys
import cv2

path = f'{os.path.dirname(sys.argv[0])}/FL-CL-ALX.xlsx'

dir = 'C:/TXT'

for dic in os.listdir('C:/'):
    if os.path.isdir(dir):
        for txt in os.listdir(dir):
            os.remove(os.path.join(dir, txt))
        print(f'Arquivos da pasta {dir} excluidos')
        break
    else:
        os.mkdir(dir)
        print(f'Pasta {dir} criada')

Tk().withdraw()
filename = askopenfilename()
print(f'Arquivo selecionado {filename}')

excel = filename

messagebox.showwarning(title= 'Aviso', message = "A partir dessa mensagem NÃO DEVERÁ MEXER NO MOUSE.\nAntes de rodar o programa, conferir se as imagens foram geradas pela tela de seu computador. Caso não tenha gerado as imagens, gerar print igual às imagens que há na pasta do programa.")

#Abrir programa pelo executar
text_file = open(f'{os.path.dirname(sys.argv[0])}\login.txt')
data= text_file.read()
usuario = data.split('\n')[0]
senha = data.split('\n')[1]

py.hotkey('winleft',  'r')        
py.write(r'P:\Sistema Comercial x64.lnk')
time.sleep(0.5) 
py.press('enter')
time.sleep(1)
#Fechar mensagem de risco 
py.hotkey('altleft',  'a')
time.sleep(0.5)
#Logar
acesso=None
x=0
taqi_png = 'taqi.png'
while acesso is None and x<60:
    #taqi = cv2.imread(f'{os.path.dirname(sys.argv[0])}\taqi.png')
    taqi=os.path.join((os.path.dirname(sys.argv[0])),taqi_png)
    acesso = py.locateOnScreen(taqi,grayscale=True, confidence=0.9)
    if acesso!= None:
        point = py.center(acesso)
        py.click(point.x,point.y)
        break
    time.sleep(0.5)
    x+=1
py.write(str(usuario))   
py.press('tab')
py.write(str(senha))
py.press('tab')
py.write('herval')
py.press('tab')
py.press('tab')
py.press('enter')
#Passos para acessar determinada área no programa
time.sleep(1.5)
py.press('altleft')
py.press('s')
py.press('e')
time.sleep(0.8)
py.press('enter')
py.press('r')
py.press('x')
py.press('enter')
time.sleep(2)
#Pegar dados do excel para escrever no programa
wb = load_workbook(excel, data_only=True)       

ws = wb.worksheets[0]

cell=2
caixa_uf = 0
for row in range(1,ws.max_row):

    fl = ws.cell(cell,1).value
    almox = ws.cell(cell,3).value
    almox2 = ws.cell(cell,4).value
    data1 = ws.cell(2,5).value
    data1 = str(data1).split(' ')
    data1 = data1[0].split('-')
    data2 = ws.cell(2,6).value
    data2 = str(data2).split(' ')
    data2 = data2[0].split('-')
    #Digita no programa
    py.write(str(fl))
    py.press('tab')
    py.write(str(fl))
    py.press('tab')
    py.write(str(almox))
    py.press('tab')
    py.write(str(almox2))

    x=0
    while x < 5:
        py.press('tab')
        x+=1

    py.write(data1[2])
    py.write(data1[1])
    py.write(data1[0])
    py.press('tab')
    py.write(data2[2])
    py.write(data2[1])
    py.write(data2[0])
    py.press('tab')

    x=0
    while x < 9:
        py.press('tab')
        x+=1

    while caixa_uf == 0:
        py.press('c')
        caixa_uf = 1
        
    py.press('tab')
    time.sleep(0.5)
    py.hotkey('altleft','f4')
    time.sleep(0.5)

    x=0
    while x < 16:
        py.press('tab')
        x+=1
    py.press('enter')
#Aguardando programa finalizar a tarefa
    print(f"Gerando relatório da filial: {fl}")
    acesso=None
    x=0
    janela_png = 'janela.png'
    while acesso is None and x<6000:
        janela = os.path.join(os.path.dirname(sys.argv[0]), janela_png)
        acesso = py.locateOnScreen(janela,grayscale=True, confidence=0.9)
        if acesso!= None:
            time.sleep(1)
            point = py.center(acesso)
            py.click(point.x,point.y)
            time.sleep(1)
            py.click(point.x,point.y)
            py.press('alt')
            py.press('a')
            py.press('c')
            py.press('x')
            py.press('enter')
            time.sleep(0.5)
            py.press('enter')
            caminho = f'C:\TXT\{str(fl)}.txt'
            time.sleep(0.5)
            py.write(caminho)
            time.sleep(0.5)
            py.press('enter')
            break
        time.sleep(1)
        x+=1
    print(f"Ainda gerando relatório da filial: {fl}")

    time.sleep(2)
    
    acesso=True
    x=0
    while acesso is not None and x<600:
        time.sleep(2)
        janela_pequena_png = 'janela_pequena.png'
        janela_pequena = os.path.join(os.path.dirname(sys.argv[0]), janela_pequena_png)  
        acesso = py.locateOnScreen(janela_pequena,grayscale=True, confidence=0.9)
        if acesso == None:
            break
        time.sleep(5)
        x+=1
    print(f"Relatório salvo: {caminho}")
    time.sleep(1)
    py.click()
    time.sleep(1)
    py.hotkey('altleft','f4')
    time.sleep(1)
    py.click()
    py.hotkey('tab')
    py.hotkey('tab')
    py.hotkey('tab')
    py.hotkey('tab')
    time.sleep(1)
    py.hotkey('altleft','f4')

    cell += 1

messagebox.showinfo(title='Informação', message='Processo ERP Concluído')